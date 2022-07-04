from datetime import datetime, timedelta, date
import imp
from time import time
import webbrowser
from wsgiref import headers
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import openpyxl
import urllib
from urllib.request import urlopen, Request



file_name = 'TimeTable/SC_CODE.xlsx'
wb = openpyxl.load_workbook(file_name, data_only=True)
sheet = wb['SC_CODE']

#전체를 함수로 만들고 main.py에서 불러오기
"""
sc_name = input("학교이름을 입력해주세요: ")

sc_code = None
row_number = 1
for row_data in sheet.iter_rows():
    #print(sheet['C{0}'.format(row_number)].value)
    if (sheet['C{0}'.format(row_number)].value == sc_name):
        #print(row_number)
        sc_code = sheet['A{}'.format(row_number)].value
    row_number += 1
if sc_code == None:
    sc_code = "학교 데이터가 없습니다."
#print(row_number)
#print(sc_code)

now = date.today()
year = now.year
to_date = None
from_date = None
#print(now)

if datetime.weekday(now) == 0:
    from_date = now
    to_date = now + timedelta(days=4)
elif datetime.weekday(now) == 1:
    from_date = now - timedelta(days=1)
    to_date = now + timedelta(days=3)
elif datetime.weekday(now) == 2:
    from_date = now - timedelta(days=2)
    to_date = now + timedelta(days=2)
elif datetime.weekday(now) == 3:
    from_date = now - timedelta(days=3)
    to_date = now + timedelta(days=1)
elif datetime.weekday(now) == 4:
    from_date = now - timedelta(days=2)
    to_date = now
elif datetime.weekday(now) == 5:
    from_date = now + timedelta(days=2)
    to_date = now + timedelta(days=6)
elif datetime.weekday(now) == 6:
    from_date = now + timedelta(days=1)
    to_date = now + timedelta(days=5)

#print(from_date)
#print(to_date)
It_from_date = str(from_date).replace("-", "")
It_to_date = str(to_date).replace("-", "")
#print(It_from_date)
#print(It_to_date)

grade = input("학년을 입력해주세요: ")
class_name = input("반을 입력해주세요: ")

API_KEY = 'dc556b41726b467e888e9306079ed091'
url = "https://open.neis.go.kr/hub/hisTimetable?KEY={0}&Type=xml&ATPT_OFCDC_SC_CODE=J10&SD_SCHUL_CODE={1}&AY={2}&GRADE={3}&CLASS_NM={4}&TI_FROM_YMD={5}&TI_TO_YMD={6}".format(API_KEY, sc_code, year, grade, class_name, It_from_date, It_to_date)
#print(url)
#hdr = {'User-Agent': 'Mozilla/5.0'}
#req = Request(url, headers=hdr)
sc_xml = urlopen(url)
bsObject = BeautifulSoup(sc_xml, "html.parser")
#print(bsObject)
data = bsObject.find_all("row")
monday = {}
tuesday = {}
wednesday = {}
thursday = {}
friday = {}
for row in data:
    #print(row)
    time_table_date = str(row.find("all_ti_ymd")).replace("<all_ti_ymd><![CDATA[", "").replace("]]></all_ti_ymd>", "")
    time_table_date_year = int(time_table_date[0:4])
    time_table_date_month = int(time_table_date[4:6])
    if time_table_date_month >= 10:
        time_table_date_month = int(str(time_table_date_month)[1])
        
    time_table_date_day = int(time_table_date[6:8])
    if time_table_date_day >= 10:
        time_table_date_day = int(str(time_table_date_day)[1])
        
    #print(time_table_date_year,time_table_date_month, time_table_date_day)
    week_day = date(time_table_date_year, time_table_date_month, time_table_date_day).weekday()
    perio = int(str(row.find("perio")).replace("<perio><![CDATA[", "").replace("]]></perio>", ""))
    itrt_cntnt = str(row.find("itrt_cntnt")).replace("<itrt_cntnt><![CDATA[", "").replace("]]></itrt_cntnt>", "")
    if week_day == 1:
        monday[perio] = {itrt_cntnt}
    elif week_day == 2:
        tuesday[perio] = {itrt_cntnt}
    elif week_day == 3:
        wednesday[perio] = {itrt_cntnt}
    elif week_day == 4:
        thursday[perio] = {itrt_cntnt}
    elif week_day == 5:
        friday[perio] = {itrt_cntnt}
        
    #print(time_table_date,perio, itrt_cntnt)

#print(monday[1], tuesday[1], wednesday[1], thursday[1], friday[1])
#print(monday.keys())
print("월요일")
for item in monday:
    print(item, "교시", str(monday[item]).replace("{'", "").replace("'}", ""))
print("화요일")
for item in tuesday:
    print(item, "교시", str(tuesday[item]).replace("{'", "").replace("'}", ""))
print("수요일")
for item in wednesday:
    print(item, "교시", str(wednesday[item]).replace("{'", "").replace("'}", ""))
print("목요일")
for item in thursday:
    print(item, "교시", str(thursday[item]).replace("{'", "").replace("'}", ""))
print("금요일")
for item in friday:
    print(item, "교시", str(friday[item]).replace("{'", "").replace("'}", ""))
"""


def TimeTable(sc_name, grade, class_name):
    #sc_name = input("학교이름을 입력해주세요: ")

    sc_code = None
    row_number = 1
    for row_data in sheet.iter_rows():
        #print(sheet['C{0}'.format(row_number)].value)
        if (sheet['C{0}'.format(row_number)].value == sc_name):
            #print(row_number)
            sc_code = sheet['A{}'.format(row_number)].value
        row_number += 1
    if sc_code == None:
        sc_code = "학교 데이터가 없습니다."
    #print(row_number)
    #print(sc_code)

    now = date.today()
    year = now.year
    to_date = None
    from_date = None
    #print(now)

    if datetime.weekday(now) == 0:
        from_date = now
        to_date = now + timedelta(days=4)
    elif datetime.weekday(now) == 1:
        from_date = now - timedelta(days=1)
        to_date = now + timedelta(days=3)
    elif datetime.weekday(now) == 2:
        from_date = now - timedelta(days=2)
        to_date = now + timedelta(days=2)
    elif datetime.weekday(now) == 3:
        from_date = now - timedelta(days=3)
        to_date = now + timedelta(days=1)
    elif datetime.weekday(now) == 4:
        from_date = now - timedelta(days=2)
        to_date = now
    elif datetime.weekday(now) == 5:
        from_date = now + timedelta(days=2)
        to_date = now + timedelta(days=6)
    elif datetime.weekday(now) == 6:
        from_date = now + timedelta(days=1)
        to_date = now + timedelta(days=5)

    #print(from_date)
    #print(to_date)
    It_from_date = str(from_date).replace("-", "")
    It_to_date = str(to_date).replace("-", "")
    #print(It_from_date)
    #print(It_to_date)

    #grade = input("학년을 입력해주세요: ")
    #class_name = input("반을 입력해주세요: ")

    API_KEY = 'dc556b41726b467e888e9306079ed091'
    url = "https://open.neis.go.kr/hub/hisTimetable?KEY={0}&Type=xml&ATPT_OFCDC_SC_CODE=J10&SD_SCHUL_CODE={1}&AY={2}&GRADE={3}&CLASS_NM={4}&TI_FROM_YMD={5}&TI_TO_YMD={6}".format(API_KEY, sc_code, year, grade, class_name, It_from_date, It_to_date)
    sc_xml = urlopen(url)
    bsObject = BeautifulSoup(sc_xml, "html.parser")
    data = bsObject.find_all("row")
    monday = {}
    tuesday = {}
    wednesday = {}
    thursday = {}
    friday = {}
    time_table = "월요일\n"
    for row in data:
        time_table_date = str(row.find("all_ti_ymd")).replace("<all_ti_ymd><![CDATA[", "").replace("]]></all_ti_ymd>", "")
        time_table_date_year = int(time_table_date[0:4])
        time_table_date_month = int(time_table_date[4:6])
        if time_table_date_month >= 10:
            time_table_date_month = int(str(time_table_date_month)[1])

        time_table_date_day = int(time_table_date[6:8])
        if time_table_date_day >= 10:
            time_table_date_day = int(str(time_table_date_day)[1])

        week_day = date(time_table_date_year, time_table_date_month, time_table_date_day).weekday()
        perio = int(str(row.find("perio")).replace("<perio><![CDATA[", "").replace("]]></perio>", ""))
        itrt_cntnt = str(row.find("itrt_cntnt")).replace("<itrt_cntnt><![CDATA[", "").replace("]]></itrt_cntnt>", "")
        if week_day == 1:
            monday[perio] = {itrt_cntnt}
        elif week_day == 2:
            tuesday[perio] = {itrt_cntnt}
        elif week_day == 3:
            wednesday[perio] = {itrt_cntnt}
        elif week_day == 4:
            thursday[perio] = {itrt_cntnt}
        elif week_day == 5:
            friday[perio] = {itrt_cntnt}

    for item in monday:
        time_table = time_table + str(item) + "교시 " + str(monday[item]).replace("{'", "").replace("'}", "") + "\n"
    time_table = time_table + "화요일\n"
    for item in tuesday:
        time_table = time_table + str(item) + "교시 " + str(tuesday[item]).replace("{'", "").replace("'}", "") + "\n"
    time_table = time_table + "수요일\n"
    for item in wednesday:
        time_table = time_table + str(item) + "교시 " + str(wednesday[item]).replace("{'", "").replace("'}", "") + "\n"
    time_table = time_table + "목요일\n"
    for item in thursday:
        time_table = time_table + str(item) + "교시 " + str(thursday[item]).replace("{'", "").replace("'}", "") + "\n"
    time_table = time_table + "금요일\n"
    for item in friday:
        time_table = time_table + str(item) + "교시 " + str(friday[item]).replace("{'", "").replace("'}", "") + "\n"
    return time_table

